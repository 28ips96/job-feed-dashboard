"""
build_excel_v3.py — Job Feed v3 Excel Builder
Adds match score columns, score breakdown, and tracking status.
"""
import json
import sys
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
from pathlib import Path
from collections import Counter

# ─── COLORS ───────────────────────────────────────────────────────────────────
NAVY       = "1C2E5A"
SLATE      = "3B5280"
PALE_BLUE  = "EEF2FB"
WHITE      = "FFFFFF"
GREEN_FILL = "E8F5E9"
SCORE_HIGH = "006400"  # 70+
SCORE_MED  = "9C6C00"  # 40-70
SCORE_LOW  = "BB0000"  # <40

CAT_STYLES = {
    "Product Manager":              ("D1ECF9", "0B5574"),
    "Technical Program Manager":    ("E8D9F7", "4B1C80"),
    "Strategy & Operations":        ("FFF3CD", "7D5A00"),
    "Finance & Strategy / FP&A":    ("D4EDDA", "155724"),
}


def thin_border():
    s = Side(style="thin", color="D0D5E8")
    return Border(left=s, right=s, top=s, bottom=s)


def days_old(posted_date):
    try:
        return (datetime.now() - datetime.strptime(posted_date, "%Y-%m-%d")).days
    except Exception:
        return None


def sort_and_filter(jobs, max_days=7):
    filtered = []
    for j in jobs:
        d = days_old(j.get("posted_date", ""))
        if d is not None and d <= max_days:
            filtered.append(j)
    # Sort by score first, then date
    return sorted(filtered,
                  key=lambda j: (j.get("match_score", 0),
                                 j.get("posted_date", "2000-01-01")),
                  reverse=True)


COLUMNS = [
    ("Score",        8),
    ("Date Posted", 12),
    ("Company",     20),
    ("Role Title",  44),
    ("Category",    22),
    ("Top Clusters", 32),
    ("Location",    28),
    ("Remote?",     10),
    ("ATS",         12),
    ("Visa",        14),
    ("Days Old",     9),
    ("MBA",          6),
    ("Status",      10),
    ("Apply",        9),
]


def write_sheet(ws, jobs, banner):
    jobs = sort_and_filter(jobs)

    # Banner row
    ws.merge_cells(f"A1:{get_column_letter(len(COLUMNS))}1")
    c = ws["A1"]
    c.value = (banner.split("·")[0].strip() +
               f"  ·  {len(jobs)} jobs  ·  US Only  ·  Mid-Senior  ·  Last 7 Days")
    c.font = Font(name="Arial", bold=True, size=13, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor=NAVY)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 36

    # Headers
    for col_i, (header, width) in enumerate(COLUMNS, 1):
        c = ws.cell(row=2, column=col_i, value=header)
        c.font = Font(name="Arial", bold=True, size=10, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor=SLATE)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = thin_border()
        ws.column_dimensions[get_column_letter(col_i)].width = width
    ws.row_dimensions[2].height = 28

    # Data rows
    for ri, job in enumerate(jobs, 3):
        cat = job.get("role_category", "")
        sponsor = job.get("visa_sponsor", "")
        d = days_old(job.get("posted_date", ""))
        is_new = d is not None and d <= 1
        score = job.get("match_score", 0)
        row_bg = GREEN_FILL if is_new else (PALE_BLUE if ri % 2 == 0 else WHITE)

        # Build top clusters string
        cluster_scores = job.get("cluster_scores", {})
        if isinstance(cluster_scores, str):
            try:
                cluster_scores = json.loads(cluster_scores)
            except Exception:
                cluster_scores = {}
        top_clusters = sorted(cluster_scores.items(), key=lambda x: x[1], reverse=True)[:3]
        cluster_str = ", ".join(f"{k.replace('_',' ').title()} ({v:.0f}%)"
                               for k, v in top_clusters if v > 0)

        status = job.get("status", "new")
        mba = "🎓" if job.get("mba_preferred") else ""

        values = [
            round(score, 1),
            job.get("posted_date", ""),
            job.get("company", ""),
            job.get("role_title", ""),
            cat,
            cluster_str,
            job.get("location", ""),
            job.get("remote", ""),
            job.get("ats_source", ""),
            sponsor,
            d if d is not None else "?",
            mba,
            status.title(),
            job.get("job_url", ""),
        ]

        for ci, val in enumerate(values, 1):
            c = ws.cell(row=ri, column=ci, value=val)
            c.fill = PatternFill("solid", fgColor=row_bg)
            c.border = thin_border()
            c.font = Font(name="Arial", size=9)
            c.alignment = Alignment(vertical="center", horizontal="left")

            # Score column
            if ci == 1:
                c.alignment = Alignment(horizontal="center", vertical="center")
                c.font = Font(name="Arial", size=10, bold=True,
                             color=SCORE_HIGH if score >= 70 else
                                   SCORE_MED if score >= 40 else SCORE_LOW)

            # Date
            if ci == 2:
                c.alignment = Alignment(horizontal="center", vertical="center")
                if is_new:
                    c.font = Font(name="Arial", size=9, bold=True, color="145A12")

            # Role title
            if ci == 4:
                c.alignment = Alignment(vertical="center", wrap_text=True)

            # Category
            if ci == 5 and cat in CAT_STYLES:
                bg, fg = CAT_STYLES[cat]
                c.fill = PatternFill("solid", fgColor=bg)
                c.font = Font(name="Arial", size=8, bold=True, color=fg)
                c.alignment = Alignment(horizontal="center", vertical="center",
                                       wrap_text=True)

            # Top clusters
            if ci == 6:
                c.font = Font(name="Arial", size=8, color="555555")
                c.alignment = Alignment(vertical="center", wrap_text=True)

            # Remote / ATS
            if ci in (8, 9):
                c.alignment = Alignment(horizontal="center", vertical="center")
                if ci == 9:
                    c.font = Font(name="Arial", size=8, color="555555")

            # Visa
            if ci == 10:
                c.alignment = Alignment(horizontal="center", vertical="center")
                if sponsor.startswith("✅"):
                    c.font = Font(name="Arial", size=9, bold=True, color="145A12")
                    c.fill = PatternFill("solid", fgColor=GREEN_FILL)
                else:
                    c.font = Font(name="Arial", size=9, color="9C6C00")

            # Days old
            if ci == 11:
                c.alignment = Alignment(horizontal="center", vertical="center")
                if isinstance(val, int):
                    if val == 0:
                        c.value = "🆕 Today"
                        c.font = Font(name="Arial", size=9, bold=True, color="006400")
                    elif val <= 3:
                        c.font = Font(name="Arial", size=9, bold=True, color="145A12")
                    elif val > 5:
                        c.font = Font(name="Arial", size=9, color="BB0000")

            # MBA badge
            if ci == 12:
                c.alignment = Alignment(horizontal="center", vertical="center")
                if val == "🎓":
                    c.font = Font(name="Arial", size=11)

            # Status
            if ci == 13:
                c.alignment = Alignment(horizontal="center", vertical="center")

            # Apply link
            if ci == 14 and val:
                c.hyperlink = val
                c.value = "→ Apply"
                c.font = Font(name="Arial", size=9, color="1A0DAB", underline="single")
                c.alignment = Alignment(horizontal="center", vertical="center")

        ws.row_dimensions[ri].height = 24

    ws.freeze_panes = "A3"
    if jobs:
        ws.auto_filter.ref = f"A2:{get_column_letter(len(COLUMNS))}{len(jobs)+2}"


def write_stats(ws, all_jobs):
    today = datetime.now().strftime("%Y-%m-%d")
    fresh = [j for j in all_jobs if (days_old(j.get("posted_date", "")) or 99) <= 7]

    ws.merge_cells("A1:C1")
    c = ws["A1"]
    c.value = f"Daily Feed Summary  ·  {datetime.now().strftime('%B %d, %Y')}"
    c.font = Font(name="Arial", bold=True, size=13, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor=NAVY)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 34

    by_cat = Counter(j.get("role_category") for j in fresh)
    visa_cnt = sum(1 for j in fresh if j.get("visa_sponsor", "").startswith("✅"))
    remote_c = sum(1 for j in fresh if "🌐" in j.get("remote", ""))
    today_c = sum(1 for j in fresh if j.get("posted_date", "") == today)
    d3_c = sum(1 for j in fresh if (days_old(j.get("posted_date", "")) or 99) <= 3)

    scored = [j for j in fresh if j.get("match_score", 0) > 0]
    avg_score = sum(j["match_score"] for j in scored) / max(len(scored), 1)
    high_score = sum(1 for j in scored if j["match_score"] >= 70)

    rows = [
        ("Metric", "Count"),
        ("Total roles (last 7 days)", len(fresh)),
        ("Posted today", today_c),
        ("Posted within 3 days", d3_c),
        ("✅ Known visa sponsors", visa_cnt),
        ("🌐 Remote-friendly", remote_c),
        ("", ""),
        ("Scoring", ""),
        ("Avg match score", f"{avg_score:.1f}"),
        ("High-fit roles (70+)", high_score),
        ("", ""),
        ("By Category", "Count"),
    ] + [(cat, cnt) for cat, cnt in by_cat.most_common()]

    for ri, (label, val) in enumerate(rows, 3):
        is_h = label in ("Metric", "By Category", "Scoring")
        bg = SLATE if is_h else (PALE_BLUE if ri % 2 == 0 else WHITE)
        fc = "FFFFFF" if is_h else "000000"
        for ci, v in enumerate([label, val], 1):
            c = ws.cell(row=ri, column=ci, value=v)
            c.font = Font(name="Arial", bold=is_h, size=10, color=fc)
            c.fill = PatternFill("solid", fgColor=bg)
            c.alignment = Alignment(horizontal="left" if ci == 1 else "center",
                                   vertical="center")
            c.border = thin_border()
        ws.row_dimensions[ri].height = 22

    # Top companies
    sep = ri + 2
    ws.merge_cells(f"A{sep}:C{sep}")
    hdr = ws[f"A{sep}"]
    hdr.value = "Top Companies (last 7 days)"
    hdr.font = Font(name="Arial", bold=True, size=11, color="FFFFFF")
    hdr.fill = PatternFill("solid", fgColor=SLATE)
    hdr.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[sep].height = 26

    co_counts = Counter(j.get("company") for j in fresh)
    for i, (co, cnt) in enumerate(co_counts.most_common(25), sep + 1):
        bg = PALE_BLUE if i % 2 == 0 else WHITE
        for ci, v in enumerate([co, cnt], 1):
            c = ws.cell(row=i, column=ci, value=v)
            c.font = Font(name="Arial", size=10)
            c.fill = PatternFill("solid", fgColor=bg)
            c.border = thin_border()
            c.alignment = Alignment(vertical="center",
                                   horizontal="left" if ci == 1 else "center")
        ws.row_dimensions[i].height = 20

    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 10


def _build_linkedin_urls_excel(company, tier, role_category="", job_title=""):
    """Build LinkedIn People Search URLs: quoted company, US geo filter, 3 degree variants."""
    import re
    import urllib.parse

    base = "https://www.linkedin.com/search/results/people/"
    qco = f'"{company}"'
    geo = "&geoUrn=%5B%22103644278%22%5D"
    origin = "&origin=GLOBAL_SEARCH_HEADER"

    rec_map = {
        "AI/ML Product Manager":    '"technical recruiter" OR "talent acquisition" AI ML',
        "Product Manager":           '"technical recruiter" OR "talent acquisition" product',
        "Technical Program Manager": '"technical recruiter" OR "talent acquisition" program',
        "Finance & Strategy / FP&A": '"recruiter" OR "talent acquisition" finance',
        "Consulting":                '"recruiter" OR "talent partner" consulting',
        "Strategy & Operations":     '"recruiter" OR "talent acquisition" strategy operations',
        "Customer Success":          '"recruiter" OR "talent acquisition" customer success',
        "Product Marketing":         '"recruiter" OR "talent acquisition" product marketing',
        "Solutions / Pre-Sales":     '"recruiter" OR "talent acquisition" solutions',
    }

    if tier == "alumni":
        kw = f'{qco} "Kelley School of Business" OR "Indiana University" MBA'
    elif tier == "recruiter":
        terms = rec_map.get(role_category, '"recruiter" OR "talent acquisition"')
        kw = f'{qco} {terms}'
    else:  # same_role
        t = (job_title or role_category or "product manager").lower()
        t = re.sub(r'^(senior|sr\.?|lead|principal|staff|associate|junior)\s+', '', t)
        t = re.sub(r'\s*[-–,|/]\s*(remote|hybrid|us|usa|sf|nyc|new york|san francisco|austin|seattle).*$',
                   '', t, flags=re.IGNORECASE)
        t = re.sub(r'\s*\(.*?\)\s*', '', t).strip()
        skip = {'the', 'and', 'for', 'with', 'at', 'in', 'of', 'a', 'an', 'to', 'or'}
        words = [w for w in t.split() if len(w) > 2 and w not in skip][:4]
        kw = f'{qco} "{" ".join(words)}"'

    enc = urllib.parse.quote_plus(kw)
    return {
        "1st": f"{base}?keywords={enc}&network=%5B%22F%22%5D{geo}{origin}",
        "2nd": f"{base}?keywords={enc}&network=%5B%22S%22%5D{geo}{origin}",
        "3rd": f"{base}?keywords={enc}&network=%5B%22O%22%5D{geo}{origin}",
    }


def write_network(ws, jobs, contacts=None):
    """Write LinkedIn networking links sheet — 3 tiers × 3 degree links per company."""
    from collections import defaultdict

    COLS   = ["Rank", "Company", "Score", "Best Role", "Tier", "1st°", "2nd°", "3rd°"]
    widths = [6,       22,        8,       38,          16,     18,      18,      18]

    ws.merge_cells(f"A1:{get_column_letter(len(COLS))}1")
    c = ws["A1"]
    c.value = "🤝 Networking Targets — LinkedIn Search Links (Alumni · Recruiters · Peers)"
    c.font = Font(name="Arial", bold=True, size=13, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor=NAVY)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 34

    for ci, (h, w) in enumerate(zip(COLS, widths), 1):
        c = ws.cell(row=2, column=ci, value=h)
        c.font = Font(name="Arial", bold=True, size=10, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor=SLATE)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = thin_border()
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[2].height = 26

    company_map = defaultdict(list)
    for j in jobs:
        company_map[j.get("company", "Unknown")].append(j)

    sorted_companies = sorted(
        company_map.items(),
        key=lambda kv: max(j.get("match_score", 0) for j in kv[1]),
        reverse=True,
    )[:50]

    TIERS = [
        ("alumni",    "E8F4FD", "👩‍🎓 Alumni"),
        ("recruiter", "FFF8E1", "🎯 Recruiters"),
        ("same_role", "F0FDF4", "💼 Peers"),
    ]

    ri = 3
    for rank, (company, cjobs) in enumerate(sorted_companies, 1):
        best     = max(j.get("match_score", 0) for j in cjobs)
        top_job  = max(cjobs, key=lambda j: j.get("match_score", 0))
        role_cat = top_job.get("role_category", "")
        role_ttl = top_job.get("role_title", "")
        sc_color = SCORE_HIGH if best >= 70 else SCORE_MED if best >= 40 else SCORE_LOW

        for t_idx, (tier, tier_bg, tier_label) in enumerate(TIERS):
            urls = _build_linkedin_urls_excel(company, tier, role_cat, role_ttl)
            row_vals = [
                rank       if t_idx == 0 else "",
                company    if t_idx == 0 else "",
                round(best, 1) if t_idx == 0 else "",
                role_ttl   if t_idx == 0 else "",
                tier_label,
                urls["1st"],
                urls["2nd"],
                urls["3rd"],
            ]
            for ci, val in enumerate(row_vals, 1):
                c = ws.cell(row=ri, column=ci, value=val)
                c.fill = PatternFill("solid", fgColor=tier_bg)
                c.border = thin_border()
                c.font = Font(name="Arial", size=9)
                c.alignment = Alignment(vertical="center")

                if ci == 1:  # Rank
                    c.alignment = Alignment(horizontal="center", vertical="center")
                    c.font = Font(name="Arial", size=9, bold=True)
                if ci == 2 and t_idx == 0:  # Company (first row of group)
                    c.font = Font(name="Arial", size=10, bold=True)
                if ci == 3 and val:  # Score
                    c.alignment = Alignment(horizontal="center", vertical="center")
                    c.font = Font(name="Arial", size=10, bold=True, color=sc_color)
                if ci == 4:  # Role title
                    c.alignment = Alignment(vertical="center", wrap_text=True)
                if ci == 5:  # Tier label
                    c.font = Font(name="Arial", size=9, bold=True)
                if ci in (6, 7, 8) and val:  # Degree links
                    deg_label  = {6: "1st°", 7: "2nd°", 8: "3rd°"}[ci]
                    link_color = "4F46E5" if ci == 6 else "374151" if ci == 7 else "9CA3AF"
                    c.hyperlink = val
                    c.value = deg_label
                    c.font = Font(name="Arial", size=9, color=link_color,
                                  underline="single", bold=(ci == 6))
                    c.alignment = Alignment(horizontal="center", vertical="center")

            ws.row_dimensions[ri].height = 20
            ri += 1

    ws.freeze_panes = "A3"


def build_excel(jobs, output_path, contacts=None):
    wb = openpyxl.Workbook()

    ws_all = wb.active
    ws_all.title = "All Jobs"
    write_sheet(ws_all, jobs, "All Roles")

    for cat, tab_name in [
        ("Product Manager",           "🚀 PM"),
        ("AI/ML Product Manager",     "🤖 AI/ML PM"),
        ("Technical Program Manager", "⚙️ TPM"),
        ("Strategy & Operations",     "📊 Strategy & Ops"),
        ("Finance & Strategy / FP&A", "💰 Finance & FP&A"),
        ("Consulting",                "🏛 Consulting"),
    ]:
        cat_jobs = [j for j in jobs if j.get("role_category") == cat]
        if cat_jobs:
            ws = wb.create_sheet(tab_name)
            write_sheet(ws, cat_jobs, cat)

    visa_jobs = [j for j in jobs if j.get("visa_sponsor", "").startswith("✅")]
    if visa_jobs:
        ws_visa = wb.create_sheet("✅ OPT-H1B Sponsors")
        write_sheet(ws_visa, visa_jobs, "Known OPT-H1B Sponsors")

    high_fit = [j for j in jobs if j.get("match_score", 0) >= 70]
    if high_fit:
        ws_fit = wb.create_sheet("🎯 High Fit")
        write_sheet(ws_fit, high_fit, "High Match Score (70+)")

    ws_net = wb.create_sheet("🤝 Network")
    write_network(ws_net, jobs, contacts=contacts)

    ws_stats = wb.create_sheet("📈 Stats")
    write_stats(ws_stats, jobs)

    wb.save(output_path)
    print(f"✅ Saved → {output_path}  ({len(wb.sheetnames)} tabs)")


if __name__ == "__main__":
    jobs_path = sys.argv[1] if len(sys.argv) > 1 else "jobs_today.json"
    out_path = sys.argv[2] if len(sys.argv) > 2 else \
        f"job_feed_{datetime.now().strftime('%Y-%m-%d')}.xlsx"
    p = Path(jobs_path)
    if not p.exists():
        print(f"❌ {jobs_path} not found")
        sys.exit(1)
    jobs = json.load(open(p))
    if not jobs:
        print("⚠️  No jobs.")
        sys.exit(0)
    build_excel(jobs, out_path)
